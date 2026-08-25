# -*- coding: utf-8 -*-
"""
Planejador de Missões Aéreas - DTA
Interface Streamlit. Toda a regra de negócio vive em regras.py; todo o dado
estático vive em dados.py. Este arquivo só orquestra e desenha.
"""

import csv
import html
import io
import json
import os
import urllib.request
import uuid
from datetime import datetime

import pandas as pd
import streamlit as st

import aba_comave

from dados import (
    AEROPORTOS,
    COORDS_BACKUP,
    DATA_ULTIMA_CONFERENCIA,
    FONTE_DADOS_AERODROMOS,
    FROTA,
)
from regras import (
    BLOQUEIO_ANAC,
    aeronaves_aptas,
    calcular_distancia_nm,
    comprimento_pista,
    decimal_para_hhmmss,
    formatar_brl,
    tempo_fixo_dta_minutos,
    velocidade_efetiva,
    verifica_restricao_pista,
)
from regras_solo import MSG_ACRESCIMO_SOLO, minutos_solo_perna

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    XLSX_DISPONIVEL = True
except ImportError:  # pragma: no cover
    XLSX_DISPONIVEL = False

# --- CONSTANTES ------------------------------------------------------------
ARQUIVO_CIDADES = "RELACAO_CIDADES_MG_ESQUILO.csv"
ARQUIVO_COORDS = "coordenadas_aerodromos.json"
URL_OURAIRPORTS = "https://davidmegginson.github.io/ourairports-data/airports.csv"

# Caixa envolvente do território brasileiro, para sanidade das coordenadas
BBOX_BRASIL = {"lat_min": -35.0, "lat_max": 6.0, "lon_min": -75.0, "lon_max": -33.0}

MSG_TOTAL_BLOQUEADO = "TOTAL INDISPONÍVEL — HÁ TRECHO NÃO OPERACIONAL (ANAC)"

st.set_page_config(page_title="Planejador de Missões - DTA", page_icon="🚁", layout="wide")

AEROPORTOS_ORDENADOS = dict(sorted(AEROPORTOS.items(), key=lambda item: item[1]["cidade"]))


# --- CARGA DE DADOS EXTERNOS ----------------------------------------------
@st.cache_data(show_spinner=False)
def carregar_cidades_extras():
    """
    Lê o catálogo de cidades sem aeródromo (áreas livres).
    Uma linha defeituosa é descartada individualmente — antes, uma linha ruim
    derrubava o arquivo inteiro.
    """
    cidades = {}
    avisos = []
    cidades_com_aeroporto = {d["cidade"] for d in AEROPORTOS.values()}

    if not os.path.exists(ARQUIVO_CIDADES):
        return {}, [
            f"Arquivo '{ARQUIVO_CIDADES}' não encontrado. "
            "Pousos em cidades sem aeroporto (Esquilo/Dauphin) ficam indisponíveis."
        ]

    try:
        with open(ARQUIVO_CIDADES, mode="r", encoding="utf-8-sig") as f:
            leitor = csv.DictReader(f)
            for num_linha, linha in enumerate(leitor, start=2):
                try:
                    cidade = str(linha["CIDADE"]).strip()
                    # aceita vírgula decimal (planilha exportada do Excel BR)
                    lat = float(str(linha["LATITUDE"]).strip().replace(",", "."))
                    lon = float(str(linha["LONGITUDE"]).strip().replace(",", "."))
                except (KeyError, TypeError, ValueError) as e:
                    avisos.append(f"Linha {num_linha} do CSV ignorada ({e}).")
                    continue

                if not cidade:
                    continue
                if not (
                    BBOX_BRASIL["lat_min"] <= lat <= BBOX_BRASIL["lat_max"]
                    and BBOX_BRASIL["lon_min"] <= lon <= BBOX_BRASIL["lon_max"]
                ):
                    avisos.append(
                        f"Linha {num_linha} ({cidade}) ignorada: coordenada fora do Brasil."
                    )
                    continue
                if cidade in cidades_com_aeroporto:
                    continue

                cidades[cidade] = {
                    "cidade": cidade,
                    "pista": "Área Livre/Campo",
                    "op_noturna": "Não",
                    "uf": "MG",
                    "lat": lat,
                    "lon": lon,
                    "is_extra": True,
                }
    except Exception as e:  # arquivo corrompido, encoding exótico etc.
        avisos.append(f"Falha ao ler '{ARQUIVO_CIDADES}': {e}")

    return dict(sorted(cidades.items())), avisos


@st.cache_data(ttl=86400, show_spinner=False)
def carregar_coordenadas():
    """
    Prioridade: arquivo local versionado > OurAirports (rede) > backup embutido.
    Devolve (coords, faltantes, origem_do_dado).
    """
    coords = dict(COORDS_BACKUP)
    origem = "backup embutido"

    if os.path.exists(ARQUIVO_COORDS):
        try:
            with open(ARQUIVO_COORDS, "r", encoding="utf-8") as f:
                locais = json.load(f)
            for icao, c in locais.items():
                if icao in AEROPORTOS:
                    coords[icao] = {"lat": float(c["lat"]), "lon": float(c["lon"])}
            origem = f"arquivo local ({ARQUIVO_COORDS})"
        except Exception:
            pass

    faltantes = [icao for icao in AEROPORTOS if icao not in coords]
    if faltantes:
        try:
            req = urllib.request.Request(URL_OURAIRPORTS, headers={"User-Agent": "DTA-Planner/1.0"})
            with urllib.request.urlopen(req, timeout=8) as resp:
                linhas = [linha.decode("utf-8") for linha in resp.readlines()]
            for linha in csv.DictReader(linhas):
                if linha["ident"] in AEROPORTOS:
                    coords[linha["ident"]] = {
                        "lat": float(linha["latitude_deg"]),
                        "lon": float(linha["longitude_deg"]),
                    }
            origem += " + OurAirports (rede)"
        except Exception:
            pass

    faltantes = sorted(icao for icao in AEROPORTOS if icao not in coords)
    return coords, faltantes, origem


CIDADES_EXTRAS, AVISOS_CIDADES = carregar_cidades_extras()
COORDS, ICAOS_SEM_COORDENADA, ORIGEM_COORDS = carregar_coordenadas()


# --- HELPERS ---------------------------------------------------------------
def obter_dados_local(local_id):
    return AEROPORTOS.get(local_id) or CIDADES_EXTRAS.get(local_id)


def obter_coordenada(local_id, dados_local):
    """Aeródromo -> tabela de coordenadas. Área livre -> o próprio registro."""
    if local_id in COORDS:
        return COORDS[local_id]["lat"], COORDS[local_id]["lon"]
    if dados_local:
        return dados_local.get("lat"), dados_local.get("lon")
    return None, None


def formatador_localidade(x):
    if x in AEROPORTOS_ORDENADOS:
        return f"{AEROPORTOS_ORDENADOS[x]['cidade']} ({x})"
    if x in CIDADES_EXTRAS:
        return f"{CIDADES_EXTRAS[x]['cidade']} (Área Livre)"
    return x


def nome_curto_aeronave(nome):
    if "Citation" in nome:
        return "Citation"
    if "Esquilo" in nome:
        return "Esquilo"
    return nome.split("(")[0].strip()


# --- ESTADO DA MISSÃO ------------------------------------------------------
# As chaves dos widgets usam o id do trecho, NUNCA o índice. Chave por índice
# faz o Streamlit reaproveitar o valor antigo quando a lista encolhe — era o
# bug que embaralhava os trechos ao clicar no ❌.
def novo_trecho(origem="SBBH", destino="SBCF", aeronave=None):
    return {
        "id": uuid.uuid4().hex[:8],
        "origem": origem,
        "destino": destino,
        "aeronave": aeronave or list(FROTA.keys())[0],
    }


def limpar_chaves(trecho):
    for prefixo in ("origem", "destino", "aeronave"):
        st.session_state.pop(f"{prefixo}_{trecho['id']}", None)


if "trechos" not in st.session_state:
    st.session_state.trechos = [novo_trecho()]
if "resultado" not in st.session_state:
    st.session_state.resultado = None


def adicionar_trecho():
    ultimo = st.session_state.trechos[-1]
    st.session_state.trechos.append(
        novo_trecho(origem=ultimo["destino"], destino="SBBH", aeronave=ultimo["aeronave"])
    )


def remover_trecho(idx):
    if len(st.session_state.trechos) > 1:
        limpar_chaves(st.session_state.trechos.pop(idx))
        st.session_state.resultado = None


def limpar_busca():
    for trecho in st.session_state.trechos:
        limpar_chaves(trecho)
    st.session_state.trechos = [novo_trecho()]
    st.session_state.resultado = None


# --- MOTOR DE CÁLCULO ------------------------------------------------------
def calcular_missao(trechos, incluir_solo=False):
    linhas = []
    erros = []
    nomes_usados = []
    custo_total = 0.0

    for i, trecho in enumerate(trechos, start=1):
        origem, destino, aeronave = trecho["origem"], trecho["destino"], trecho["aeronave"]
        dados_aero = FROTA[aeronave]

        nome_curto = nome_curto_aeronave(aeronave)
        if nome_curto not in nomes_usados:
            nomes_usados.append(nome_curto)

        dados_origem = obter_dados_local(origem)
        dados_destino = obter_dados_local(destino)
        if dados_origem is None or dados_destino is None:
            faltando = origem if dados_origem is None else destino
            erros.append(f"Trecho {i}: localidade '{faltando}' não encontrada na base.")
            continue

        bloq_o, msg_o, tipo_o = verifica_restricao_pista(origem, aeronave, dados_aero, dados_origem)
        bloq_d, msg_d, tipo_d = verifica_restricao_pista(destino, aeronave, dados_aero, dados_destino)

        bloqueio_anac = tipo_o == BLOQUEIO_ANAC or tipo_d == BLOQUEIO_ANAC
        if bloq_o and tipo_o != BLOQUEIO_ANAC:
            erros.append(f"Trecho {i}: {msg_o}")
        if bloq_d and tipo_d != BLOQUEIO_ANAC:
            erros.append(f"Trecho {i}: {msg_d}")

        lat_o, lon_o = obter_coordenada(origem, dados_origem)
        lat_d, lon_d = obter_coordenada(destino, dados_destino)

        if None in (lat_o, lon_o, lat_d, lon_d):
            dist_nm = 0.0
            if not bloqueio_anac:
                erros.append(f"Trecho {i}: coordenadas indisponíveis para {origem} → {destino}.")
        else:
            dist_nm = calcular_distancia_nm(lat_o, lon_o, lat_d, lon_d)

        if origem == destino:
            erros.append(f"Trecho {i}: origem e destino iguais ({origem}).")

        # Única exceção ao cálculo por distância na tabela DTA: pares com
        # tempo tabelado (BH x Confins = 10 min), faturados por esse tempo.
        minutos_fixos = tempo_fixo_dta_minutos(origem, destino)
        if minutos_fixos is not None:
            vel_kt = 0
            tempo_decimal = minutos_fixos / 60
        else:
            vel_kt = velocidade_efetiva(dados_aero, dist_nm)
            tempo_decimal = dist_nm / vel_kt if vel_kt > 0 else 0.0
        custo_perna = tempo_decimal * dados_aero["valor_hora"]

        # Acréscimo de solo opcional na DTA: entra APENAS no tempo exibido,
        # nunca no custo. Trecho de tempo tabelado não recebe acréscimo.
        minutos_solo = 0
        if incluir_solo and minutos_fixos is None:
            minutos_solo = minutos_solo_perna(
                origem, destino, dados_aero.get("is_heli", False)
            )
        tempo_exibido = tempo_decimal + minutos_solo / 60

        if bloqueio_anac:
            tempo_txt = "NÃO OPERACIONAL (ANAC)"
            custo_txt = "NÃO OPERACIONAL (ANAC)"
            custo_perna = 0.0
        else:
            tempo_txt = decimal_para_hhmmss(tempo_exibido)
            custo_txt = formatar_brl(custo_perna)
            custo_total += custo_perna

        linhas.append(
            {
                "mun_orig": dados_origem["cidade"],
                "uf_orig": dados_origem["uf"],
                "ind_orig": origem if origem in AEROPORTOS else "Área Livre",
                "mun_dest": dados_destino["cidade"],
                "uf_dest": dados_destino["uf"],
                "ind_dest": destino if destino in AEROPORTOS else "Área Livre",
                "anv": dados_aero["tipo_sigla"],
                "dist_nm": dist_nm,
                "vel_kt": vel_kt,
                "tempo_fixo": minutos_fixos is not None,
                "minutos_solo": minutos_solo,
                "tempo_voo_txt": decimal_para_hhmmss(tempo_decimal),
                "tempo_txt": tempo_txt,
                "custo_txt": custo_txt,
                "custo_num": custo_perna,
                "pax": dados_aero["pax"],
                "bloqueado": bloqueio_anac,
            }
        )

    if len(nomes_usados) > 1:
        titulo_frota = f"AERONAVES: {' + '.join(nomes_usados)}".upper()
    elif nomes_usados:
        titulo_frota = f"AERONAVE: {nomes_usados[0]}".upper()
    else:
        titulo_frota = "AERONAVE"

    total_bloqueado = any(linha["bloqueado"] for linha in linhas)

    return {
        "linhas": linhas,
        "erros": erros,
        "titulo_frota": titulo_frota,
        "custo_total": custo_total,
        # Decisão de negócio: havendo trecho embargado pela ANAC, o total NÃO
        # é divulgado. O cliente vê a tabela, identifica o problema e refaz o
        # planejamento antes de ter um número para levar adiante.
        "total_bloqueado": total_bloqueado,
        "incluir_solo": incluir_solo,
        "gerado_em": datetime.now(),
    }


# --- RENDERIZAÇÃO HTML -----------------------------------------------------
def _td(conteudo, extra=""):
    return f"<td style='padding: 8px; border: 1px solid #000000;{extra}'>{conteudo}</td>"


def montar_tabela_html(resultado):
    """
    Todo conteúdo dinâmico passa por html.escape(). O nome da cidade vem de um
    CSV editável por terceiros; sem escape, uma linha maliciosa injetaria
    markup direto na página (XSS - Cross-Site Scripting / execução de script
    entre sites).
    """
    vermelho = "color: #FF0000; font-weight: bold;"
    corpo = ""
    for linha in resultado["linhas"]:
        pax = html.escape(str(linha["pax"])).replace("\n", "<br>")
        if linha["bloqueado"]:
            tempo = f"<span style='{vermelho}'>{html.escape(linha['tempo_txt'])}</span>"
            custo = f"<span style='{vermelho}'>{html.escape(linha['custo_txt'])}</span>"
        else:
            tempo = html.escape(linha["tempo_txt"])
            custo = html.escape(linha["custo_txt"])

        corpo += (
            "<tr style='background-color: #ffffff; color: #000000; text-align: center;'>"
            + _td(html.escape(str(linha["mun_orig"])))
            + _td(html.escape(str(linha["uf_orig"])))
            + _td(html.escape(str(linha["ind_orig"])))
            + _td(html.escape(str(linha["mun_dest"])))
            + _td(html.escape(str(linha["uf_dest"])))
            + _td(html.escape(str(linha["ind_dest"])))
            + _td(html.escape(str(linha["anv"])), " font-weight: bold; background-color: #f9f9f9;")
            + _td(tempo)
            + _td(custo)
            + _td(pax)
            + "</tr>"
        )

    if resultado["total_bloqueado"]:
        celula_total = (
            f"<td colspan='2' style='padding: 10px; text-align: left; "
            f"border: 1px solid #000000; {vermelho}'>{MSG_TOTAL_BLOQUEADO}</td>"
        )
    else:
        celula_total = (
            f"<td colspan='2' style='padding: 10px; text-align: left; "
            f"border: 1px solid #000000;'>{formatar_brl(resultado['custo_total'])}</td>"
        )

    linha_nota = ""
    if resultado.get("incluir_solo"):
        linha_nota = (
            "<tr style='background-color: #ffffff; color: #000000; font-size: 11px;'>"
            f"<td colspan='10' style='padding: 6px; border: 1px solid #000000; "
            f"text-align: left;'>{html.escape(MSG_ACRESCIMO_SOLO)}</td></tr>"
        )

    cabecalhos = [
        "MUNICÍPIO - DECOLAGEM", "UF", "ICAO",
        "MUNICÍPIO - POUSO", "UF", "ICAO",
        "ANV", "TEMPO DE<br>DESLOCAMENTO",
        "CUSTO TOTAL<br>ESTIMADO DA MISSÃO", "CAPACIDADE",
    ]
    linha_cabecalho = "".join(
        f"<th style='padding: 8px; border: 1px solid #000000;'>{c}</th>" for c in cabecalhos
    )

    return (
        "<div style='overflow-x: auto;'>"
        "<table style='width:100%; text-align:center; border-collapse: collapse; "
        "font-family: sans-serif; border: 2px solid #000000;'>"
        "<thead>"
        "<tr style='background-color: #9bc2e6; color: #000000; border: 1px solid #000000;'>"
        "<th colspan='3' style='padding: 8px; border: 1px solid #000000;'>ORIGEM</th>"
        "<th colspan='3' style='padding: 8px; border: 1px solid #000000;'>DESTINO</th>"
        f"<th colspan='4' style='padding: 8px; border: 1px solid #000000;'>"
        f"{html.escape(resultado['titulo_frota'])}</th>"
        "</tr>"
        f"<tr style='background-color: #ddebf7; color: #000000; font-size: 13px;'>{linha_cabecalho}</tr>"
        "</thead><tbody>"
        f"{corpo}"
        "<tr style='background-color: #f2f2f2; color: #000000; font-weight: bold;'>"
        "<td colspan='8' style='padding: 10px; text-align: right; border: 1px solid #000000;'>TOTAL</td>"
        f"{celula_total}"
        "</tr>"
        f"{linha_nota}"
        "</tbody></table></div>"
    )


# --- EXPORTAÇÃO XLSX -------------------------------------------------------
def gerar_xlsx_missao(resultado):
    """Planilha com o mesmo layout da tabela, custos como número (somável)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Missão"

    borda = Border(*[Side(style="thin", color="000000")] * 4)
    fill_1 = PatternFill("solid", fgColor="9BC2E6")
    fill_2 = PatternFill("solid", fgColor="DDEBF7")
    fill_total = PatternFill("solid", fgColor="F2F2F2")
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    negrito = Font(bold=True)
    vermelho = Font(bold=True, color="FF0000")

    ws.merge_cells("A1:C1")
    ws.merge_cells("D1:F1")
    ws.merge_cells("G1:J1")
    ws["A1"], ws["D1"], ws["G1"] = "ORIGEM", "DESTINO", resultado["titulo_frota"]

    cabecalhos = [
        "MUNICÍPIO - DECOLAGEM", "UF", "ICAO",
        "MUNICÍPIO - POUSO", "UF", "ICAO",
        "ANV", "TEMPO DE DESLOCAMENTO",
        "CUSTO ESTIMADO", "CAPACIDADE",
    ]
    ws.append(cabecalhos)

    for linha in resultado["linhas"]:
        ws.append([
            linha["mun_orig"], linha["uf_orig"], linha["ind_orig"],
            linha["mun_dest"], linha["uf_dest"], linha["ind_dest"],
            linha["anv"], linha["tempo_txt"],
            linha["custo_txt"] if linha["bloqueado"] else linha["custo_num"],
            linha["pax"],
        ])

    linha_total = ws.max_row + 1
    ws.merge_cells(start_row=linha_total, start_column=1, end_row=linha_total, end_column=8)
    ws.cell(row=linha_total, column=1, value="TOTAL")
    ws.merge_cells(start_row=linha_total, start_column=9, end_row=linha_total, end_column=10)
    celula_total = ws.cell(row=linha_total, column=9)
    if resultado["total_bloqueado"]:
        celula_total.value = MSG_TOTAL_BLOQUEADO
        celula_total.font = vermelho
    else:
        celula_total.value = resultado["custo_total"]
        celula_total.number_format = 'R$ #,##0.00'

    if resultado.get("incluir_solo"):
        linha_obs = linha_total + 1
        ws.merge_cells(start_row=linha_obs, start_column=1, end_row=linha_obs, end_column=10)
        celula_obs = ws.cell(row=linha_obs, column=1, value=MSG_ACRESCIMO_SOLO)
        celula_obs.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        celula_obs.font = Font(size=9, italic=True)

    # formatação geral
    for linha_cels in ws.iter_rows(min_row=1, max_row=linha_total, max_col=10):
        for celula in linha_cels:
            celula.border = borda
            celula.alignment = centro

    for col in range(1, 11):
        ws.cell(row=1, column=col).fill = fill_1
        ws.cell(row=1, column=col).font = negrito
        ws.cell(row=2, column=col).fill = fill_2
        ws.cell(row=2, column=col).font = negrito
        ws.cell(row=linha_total, column=col).fill = fill_total
        ws.cell(row=linha_total, column=col).font = negrito

    for idx_linha, linha in enumerate(resultado["linhas"], start=3):
        celula_tempo = ws.cell(row=idx_linha, column=8)
        celula_custo = ws.cell(row=idx_linha, column=9)
        if linha["bloqueado"]:
            celula_tempo.font = vermelho
            celula_custo.font = vermelho
        else:
            celula_custo.number_format = 'R$ #,##0.00'

    larguras = [26, 6, 12, 26, 6, 12, 14, 20, 22, 24]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_xlsx_aerodromos(df):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Aeródromos")
        ws = writer.sheets["Aeródromos"]
        for i, coluna in enumerate(df.columns, start=1):
            valores = df[coluna].astype(str)
            largura = max(len(str(coluna)), int(valores.str.len().max() or 0)) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(largura, 40)
    return buffer.getvalue()


# ===========================================================================
# INTERFACE
# ===========================================================================
st.title("🚁 Planejador de Missões Aéreas - DTA")

aba_missao, aba_comave_ui, aba_consulta = st.tabs(
    ["✈️ Cálculo de Missão (DTA)", "🛡️ Cálculo COMAVE", "🛬 Consulta de Aeródromos"]
)

# --------------------------- ABA 1: MISSÃO ---------------------------------
with aba_missao:
    st.caption("Roteirizador dinâmico point-to-point com especificação de aeronave por perna.")

    for aviso in AVISOS_CIDADES:
        st.info(f"💡 {aviso}")
    if ICAOS_SEM_COORDENADA:
        st.warning(
            f"⚠️ {len(ICAOS_SEM_COORDENADA)} aeródromo(s) sem coordenada: "
            f"{', '.join(ICAOS_SEM_COORDENADA)}. Rotas envolvendo esses locais não serão "
            f"calculadas. Rode `python gerar_coordenadas.py` e versione o arquivo "
            f"`{ARQUIVO_COORDS}` no repositório."
        )

    st.divider()

    incluir_solo = st.checkbox(
        "Incluir acréscimo de tempo de solo (15 min por aeroporto, 25 min se capital, "
        "5 min por perna de helicóptero)",
        key="dta_incluir_solo",
        help="Afeta apenas o tempo exibido. O custo continua calculado pelo tempo de voo.",
    )

    st.divider()

    for i, trecho in enumerate(st.session_state.trechos):
        tid = trecho["id"]
        if i > 0:
            st.markdown(
                "<hr style='border: 2px solid #00008B; margin-top: 15px; margin-bottom: 25px;'>",
                unsafe_allow_html=True,
            )
        st.markdown(f"**TRECHO {i + 1}**")

        col_aero, col_orig, col_dest, col_del = st.columns([3.5, 3, 3, 0.5])

        with col_aero:
            aeronave = st.selectbox(
                "Aeronave (Equipamento)",
                options=list(FROTA.keys()),
                index=list(FROTA.keys()).index(trecho["aeronave"]),
                key=f"aeronave_{tid}",
            )
            trecho["aeronave"] = aeronave

        opcoes = list(AEROPORTOS_ORDENADOS.keys())
        if FROTA[aeronave].get("is_heli"):
            opcoes += list(CIDADES_EXTRAS.keys())

        # Se o usuário trocar de helicóptero para asa fixa, a área livre some da
        # lista. É obrigatório apagar a chave do widget junto, senão o Streamlit
        # levanta exceção por valor fora das opções.
        for campo in ("origem", "destino"):
            if trecho[campo] not in opcoes:
                trecho[campo] = opcoes[0]
                st.session_state.pop(f"{campo}_{tid}", None)

        with col_orig:
            trecho["origem"] = st.selectbox(
                "Origem",
                options=opcoes,
                index=opcoes.index(trecho["origem"]),
                format_func=formatador_localidade,
                key=f"origem_{tid}",
            )

        with col_dest:
            trecho["destino"] = st.selectbox(
                "Destino",
                options=opcoes,
                index=opcoes.index(trecho["destino"]),
                format_func=formatador_localidade,
                key=f"destino_{tid}",
            )

        with col_del:
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            if len(st.session_state.trechos) > 1:
                st.button(
                    "❌", key=f"del_{tid}", on_click=remover_trecho, args=(i,),
                    help="Remover este trecho",
                )

    st.write("")
    col_add, _, col_clear = st.columns([2, 8, 2])
    col_add.button("➕ Adicionar Trecho", on_click=adicionar_trecho)
    col_clear.button("🗑️ Limpar Busca", on_click=limpar_busca, type="secondary",
                     width="stretch")

    st.divider()

    if st.button("Calcular Missão Completa", type="primary", width="stretch"):
        st.session_state.resultado = calcular_missao(
            st.session_state.trechos, incluir_solo=incluir_solo
        )

    # O resultado é renderizado FORA do if do botão. Clicar em "Baixar XLSX"
    # dispara um rerun do script; se a tabela dependesse do botão de cálculo,
    # ela desapareceria da tela no exato momento do download.
    resultado = st.session_state.resultado
    if resultado:
        if resultado["erros"]:
            for erro in resultado["erros"]:
                st.error(f"🚨 {erro}")
        elif resultado["linhas"]:
            if resultado["total_bloqueado"]:
                st.warning(
                    "⚠️ Há trecho não operacional pela ANAC (Agência Nacional de Aviação Civil). "
                    "A tabela foi gerada, mas o custo total está bloqueado até que o "
                    "planejamento seja ajustado."
                )
            else:
                st.success("✅ Rotas calculadas e validadas com sucesso.")

            if XLSX_DISPONIVEL:
                carimbo = resultado["gerado_em"].strftime("%Y%m%d_%H%M")
                st.download_button(
                    "⬇️ Baixar tabela em Excel (.xlsx)",
                    data=gerar_xlsx_missao(resultado),
                    file_name=f"missao_dta_{carimbo}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.warning("Exportação indisponível: instale a biblioteca `openpyxl`.")

            st.markdown(montar_tabela_html(resultado), unsafe_allow_html=True)

            with st.expander("🔍 Memória de cálculo (conferência)"):
                st.dataframe(
                    pd.DataFrame([
                        {
                            "Trecho": f"{l['ind_orig']} → {l['ind_dest']}",
                            "ANV": l["anv"],
                            "Distância (NM)": l["dist_nm"],
                            "Velocidade (KT)": "Tabelado" if l.get("tempo_fixo") else str(l["vel_kt"]),
                            "Tempo de voo": l["tempo_voo_txt"],
                            "Solo (min)": l["minutos_solo"],
                            "Tempo exibido": l["tempo_txt"],
                            "Custo (R$)": round(l["custo_num"], 2),
                        }
                        for l in resultado["linhas"]
                    ]),
                    width="stretch",
                    hide_index=True,
                )
                st.caption(
                    "Distância ortodrômica (círculo máximo) entre aeródromos. "
                    "BH x Confins usa tempo tabelado de 10 minutos, já faturado. "
                    "O acréscimo de solo, quando marcado, entra apenas no tempo "
                    "exibido — o custo sai sempre do tempo de voo."
                )

# --------------------------- ABA 2: COMAVE ---------------------------------
with aba_comave_ui:
    aba_comave.render(
        AEROPORTOS_ORDENADOS,
        CIDADES_EXTRAS,
        formatador_localidade,
        obter_dados_local,
        obter_coordenada,
    )

# -------------------------- ABA 3: CONSULTA --------------------------------
with aba_consulta:
    st.caption("Consulta dos aeródromos cadastrados. Áreas livres não entram nesta relação.")

    registros = []
    for icao, dados in AEROPORTOS_ORDENADOS.items():
        comprimento = comprimento_pista(dados["pista"])
        coord = COORDS.get(icao)
        registros.append(
            {
                "ICAO": icao,
                "Município": dados["cidade"],
                "UF": dados["uf"],
                "Pista": dados["pista"],
                # None (e não "—"): coluna com tipos misturados quebra a
                # serialização Arrow usada pelo st.dataframe
                "Comprimento (m)": comprimento,
                "Operação noturna": dados["op_noturna"],
                "Restrição ANAC": "Sim" if dados.get("restricao_anac") else "Não",
                "Latitude": round(coord["lat"], 4) if coord else None,
                "Longitude": round(coord["lon"], 4) if coord else None,
            }
        )
    df_aerodromos = pd.DataFrame(registros)
    df_aerodromos["Comprimento (m)"] = df_aerodromos["Comprimento (m)"].astype("Int64")

    col_busca, col_uf = st.columns([3, 2])
    with col_busca:
        termo = st.text_input("Buscar por município ou código ICAO", key="busca_aerodromo")
    with col_uf:
        ufs = st.multiselect(
            "Filtrar por UF", options=sorted(df_aerodromos["UF"].unique()), key="filtro_uf"
        )

    df_filtrado = df_aerodromos
    if termo:
        alvo = termo.strip().upper()
        df_filtrado = df_filtrado[
            df_filtrado["Município"].str.upper().str.contains(alvo, na=False)
            | df_filtrado["ICAO"].str.upper().str.contains(alvo, na=False)
        ]
    if ufs:
        df_filtrado = df_filtrado[df_filtrado["UF"].isin(ufs)]

    if df_filtrado.empty:
        st.info("Nenhum aeródromo corresponde ao filtro. Ajuste a busca ou a UF.")
    else:
        icaos_filtrados = list(df_filtrado["ICAO"])
        selecionado = st.selectbox(
            "Ficha do aeródromo",
            options=icaos_filtrados,
            format_func=lambda x: f"{AEROPORTOS[x]['cidade']} ({x}) - {AEROPORTOS[x]['uf']}",
            key="ficha_aerodromo",
        )
        dados_sel = AEROPORTOS[selecionado]

        st.markdown(f"### {dados_sel['cidade']} — {selecionado} / {dados_sel['uf']}")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Pista", dados_sel["pista"])
        c2.metric("Operação noturna", dados_sel["op_noturna"])
        c3.metric("Restrição ANAC", "Sim" if dados_sel.get("restricao_anac") else "Não")
        coord_sel = COORDS.get(selecionado)
        c4.metric(
            "Coordenadas",
            f"{coord_sel['lat']:.3f}, {coord_sel['lon']:.3f}" if coord_sel else "Indisponível",
        )

        aptas = aeronaves_aptas(selecionado, dados_sel, FROTA)
        inaptas = [nome for nome in FROTA if nome not in aptas]
        col_ok, col_nok = st.columns(2)
        with col_ok:
            st.markdown("**Aeronaves aptas**")
            st.write("\n".join(f"- {n}" for n in aptas) if aptas else "Nenhuma.")
        with col_nok:
            st.markdown("**Aeronaves impedidas**")
            st.write("\n".join(f"- {n}" for n in inaptas) if inaptas else "Nenhuma.")

        st.caption(
            "Aptidão baseada apenas em restrição regulatória e comprimento mínimo de pista. "
            "Não considera peso, carga, altitude-densidade nem condições meteorológicas."
        )

        st.divider()
        st.markdown(f"**Relação completa** ({len(df_filtrado)} aeródromo(s))")
        st.dataframe(df_filtrado, width="stretch", hide_index=True)

        if XLSX_DISPONIVEL:
            st.download_button(
                "⬇️ Baixar relação em Excel (.xlsx)",
                data=gerar_xlsx_aerodromos(df_filtrado),
                file_name=f"aerodromos_dta_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_aerodromos",
            )

    st.caption(
        f"Fonte dos dados de pista/balizamento: {FONTE_DADOS_AERODROMOS} · "
        f"Última conferência: {DATA_ULTIMA_CONFERENCIA} · "
        f"Coordenadas: {ORIGEM_COORDS}."
    )
