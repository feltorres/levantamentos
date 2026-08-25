# -*- coding: utf-8 -*-
"""
Aba de cálculo do COMAVE.

Isolada do fluxo da DTA: estado próprio no session_state, tabela de preços
própria e regras de tempo próprias. Nenhuma função daqui é chamada pela aba
DTA — o que compartilham é apenas geografia (aeródromos e coordenadas) e as
validações operacionais de pista/ANAC, que são fatos do mundo, não preço.
"""

import html
import io
import uuid
from datetime import datetime

import pandas as pd
import streamlit as st

from dados import AEROPORTOS
from dados_comave import CONTRATOS, FROTA_COMAVE, REFERENCIA_NOTA_TECNICA
from regras import (
    BLOQUEIO_ANAC,
    calcular_distancia_nm,
    decimal_para_hhmmss,
    formatar_brl,
    verifica_restricao_pista,
)
from regras_comave import (
    aeronaves_do_contrato,
    calcular_tempos_comave,
    valor_hora,
)
from regras_solo import MSG_ACRESCIMO_SOLO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    XLSX_DISPONIVEL = True
except ImportError:  # pragma: no cover
    XLSX_DISPONIVEL = False

MSG_TOTAL_BLOQUEADO = "TOTAL INDISPONÍVEL — HÁ TRECHO NÃO OPERACIONAL (ANAC)"
NOTA_SOLO = MSG_ACRESCIMO_SOLO


# --- ESTADO ----------------------------------------------------------------
def _novo_trecho(aeronave, origem="SBBH", destino="SBCF"):
    return {"id": uuid.uuid4().hex[:8], "origem": origem, "destino": destino,
            "aeronave": aeronave}


def _limpar_chaves(trecho):
    for prefixo in ("origem", "destino", "aeronave"):
        st.session_state.pop(f"cmv_{prefixo}_{trecho['id']}", None)


def _remover_trecho(idx):
    if len(st.session_state.cmv_trechos) > 1:
        _limpar_chaves(st.session_state.cmv_trechos.pop(idx))
        st.session_state.cmv_resultado = None


def _adicionar_trecho():
    ultimo = st.session_state.cmv_trechos[-1]
    st.session_state.cmv_trechos.append(
        _novo_trecho(ultimo["aeronave"], origem=ultimo["destino"], destino="SBBH")
    )


def _limpar(aeronave_padrao):
    for trecho in st.session_state.cmv_trechos:
        _limpar_chaves(trecho)
    st.session_state.cmv_trechos = [_novo_trecho(aeronave_padrao)]
    st.session_state.cmv_resultado = None


# --- CÁLCULO ---------------------------------------------------------------
def calcular_missao_comave(trechos, contrato, nome_contrato, obter_dados_local, obter_coordenada):
    linhas, erros, nomes_usados = [], [], []
    custo_total = 0.0

    for i, trecho in enumerate(trechos, start=1):
        origem, destino, aeronave = trecho["origem"], trecho["destino"], trecho["aeronave"]
        dados_aero = FROTA_COMAVE[aeronave]

        preco_hora = valor_hora(contrato, aeronave)
        if preco_hora is None:
            erros.append(
                f"Trecho {i}: {aeronave} não está coberta pelo contrato {nome_contrato}."
            )
            continue

        nome_curto = "Citation" if "Citation" in aeronave else aeronave.split("(")[0].strip()
        if nome_curto not in nomes_usados:
            nomes_usados.append(nome_curto)

        dados_origem = obter_dados_local(origem)
        dados_destino = obter_dados_local(destino)
        if dados_origem is None or dados_destino is None:
            faltando = origem if dados_origem is None else destino
            erros.append(f"Trecho {i}: localidade '{faltando}' não encontrada na base.")
            continue

        bloq_o, msg_o, tipo_o = verifica_restricao_pista(origem, aeronave, dados_aero, dados_origem)
        bloq_d, msg_d, tipo_d = verifica_restricao_pista(destino, aeronave, dados_aero, dados_destino)
        bloqueio_anac = tipo_o == BLOQUEIO_ANAC or tipo_d == BLOQUEIO_ANAC
        if bloq_o and tipo_o != BLOQUEIO_ANAC:
            erros.append(f"Trecho {i}: {msg_o}")
        if bloq_d and tipo_d != BLOQUEIO_ANAC:
            erros.append(f"Trecho {i}: {msg_d}")

        lat_o, lon_o = obter_coordenada(origem, dados_origem)
        lat_d, lon_d = obter_coordenada(destino, dados_destino)
        if None in (lat_o, lon_o, lat_d, lon_d):
            dist_nm = 0.0
            if not bloqueio_anac:
                erros.append(f"Trecho {i}: coordenadas indisponíveis para {origem} → {destino}.")
        else:
            dist_nm = calcular_distancia_nm(lat_o, lon_o, lat_d, lon_d)

        if origem == destino:
            erros.append(f"Trecho {i}: origem e destino iguais ({origem}).")

        tempos = calcular_tempos_comave(origem, destino, dados_aero, dist_nm)
        custo_perna = tempos["horas_faturadas"] * preco_hora

        if bloqueio_anac:
            tempo_txt = custo_txt = "NÃO OPERACIONAL (ANAC)"
            custo_perna = 0.0
        else:
            tempo_txt = decimal_para_hhmmss(tempos["tempo_total_h"])
            custo_txt = formatar_brl(custo_perna)
            custo_total += custo_perna

        linhas.append({
            "mun_orig": dados_origem["cidade"],
            "uf_orig": dados_origem["uf"],
            "ind_orig": origem if origem in AEROPORTOS else "Área Livre",
            "mun_dest": dados_destino["cidade"],
            "uf_dest": dados_destino["uf"],
            "ind_dest": destino if destino in AEROPORTOS else "Área Livre",
            "anv": dados_aero["tipo_sigla"],
            "dist_nm": dist_nm,
            "vel_kt": tempos["vel_kt"],
            "tempo_voo_txt": decimal_para_hhmmss(tempos["tempo_voo_h"]),
            "minutos_solo": tempos["minutos_solo"],
            "tempo_fixo": tempos["tempo_fixo"],
            "horas_faturadas": tempos["horas_faturadas"],
            "valor_hora": preco_hora,
            "tempo_txt": tempo_txt,
            "custo_txt": custo_txt,
            "custo_num": custo_perna,
            "pax": dados_aero["pax"],
            "bloqueado": bloqueio_anac,
        })

    if len(nomes_usados) > 1:
        titulo_frota = f"AERONAVES: {' + '.join(nomes_usados)}".upper()
    elif nomes_usados:
        titulo_frota = f"AERONAVE: {nomes_usados[0]}".upper()
    else:
        titulo_frota = "AERONAVE"

    return {
        "linhas": linhas,
        "erros": erros,
        "titulo_frota": titulo_frota,
        "custo_total": custo_total,
        "total_bloqueado": any(linha["bloqueado"] for linha in linhas),
        "contrato": nome_contrato,
        "sei": contrato["sei"],
        "vigencia": contrato["vigencia"],
        "gerado_em": datetime.now(),
    }


# --- TABELA HTML -----------------------------------------------------------
def _td(conteudo, extra=""):
    return f"<td style='padding: 8px; border: 1px solid #000000;{extra}'>{conteudo}</td>"


def montar_tabela_html(resultado):
    vermelho = "color: #FF0000; font-weight: bold;"
    corpo = ""
    for linha in resultado["linhas"]:
        pax = html.escape(str(linha["pax"])).replace("\n", "<br>")
        if linha["bloqueado"]:
            tempo = f"<span style='{vermelho}'>{html.escape(linha['tempo_txt'])}</span>"
            custo = f"<span style='{vermelho}'>{html.escape(linha['custo_txt'])}</span>"
        else:
            tempo = html.escape(linha["tempo_txt"])
            custo = html.escape(linha["custo_txt"])
        corpo += (
            "<tr style='background-color: #ffffff; color: #000000; text-align: center;'>"
            + _td(html.escape(str(linha["mun_orig"])))
            + _td(html.escape(str(linha["uf_orig"])))
            + _td(html.escape(str(linha["ind_orig"])))
            + _td(html.escape(str(linha["mun_dest"])))
            + _td(html.escape(str(linha["uf_dest"])))
            + _td(html.escape(str(linha["ind_dest"])))
            + _td(html.escape(str(linha["anv"])), " font-weight: bold; background-color: #f9f9f9;")
            + _td(tempo)
            + _td(custo)
            + _td(pax)
            + "</tr>"
        )

    if resultado["total_bloqueado"]:
        celula_total = (
            f"<td colspan='2' style='padding: 10px; text-align: left; "
            f"border: 1px solid #000000; {vermelho}'>{MSG_TOTAL_BLOQUEADO}</td>"
        )
    else:
        celula_total = (
            f"<td colspan='2' style='padding: 10px; text-align: left; "
            f"border: 1px solid #000000;'>{formatar_brl(resultado['custo_total'])}</td>"
        )

    cabecalhos = [
        "MUNICÍPIO - DECOLAGEM", "UF", "ICAO", "MUNICÍPIO - POUSO", "UF", "ICAO",
        "ANV", "TEMPO DE<br>DESLOCAMENTO", "CUSTO TOTAL<br>ESTIMADO DA MISSÃO", "CAPACIDADE",
    ]
    linha_cabecalho = "".join(
        f"<th style='padding: 8px; border: 1px solid #000000;'>{c}</th>" for c in cabecalhos
    )
    subtitulo = html.escape(f"{resultado['contrato']} · SEI {resultado['sei']}")

    return (
        "<div style='overflow-x: auto;'>"
        "<table style='width:100%; text-align:center; border-collapse: collapse; "
        "font-family: sans-serif; border: 2px solid #000000;'>"
        "<thead>"
        "<tr style='background-color: #c6e0b4; color: #000000;'>"
        f"<th colspan='10' style='padding: 8px; border: 1px solid #000000;'>{subtitulo}</th>"
        "</tr>"
        "<tr style='background-color: #9bc2e6; color: #000000; border: 1px solid #000000;'>"
        "<th colspan='3' style='padding: 8px; border: 1px solid #000000;'>ORIGEM</th>"
        "<th colspan='3' style='padding: 8px; border: 1px solid #000000;'>DESTINO</th>"
        f"<th colspan='4' style='padding: 8px; border: 1px solid #000000;'>"
        f"{html.escape(resultado['titulo_frota'])}</th>"
        "</tr>"
        f"<tr style='background-color: #ddebf7; color: #000000; font-size: 13px;'>{linha_cabecalho}</tr>"
        "</thead><tbody>"
        f"{corpo}"
        "<tr style='background-color: #f2f2f2; color: #000000; font-weight: bold;'>"
        "<td colspan='8' style='padding: 10px; text-align: right; border: 1px solid #000000;'>TOTAL</td>"
        f"{celula_total}"
        "</tr>"
        "<tr style='background-color: #ffffff; color: #000000; font-size: 11px;'>"
        f"<td colspan='10' style='padding: 6px; border: 1px solid #000000; text-align: left;'>"
        f"{html.escape(NOTA_SOLO)}</td>"
        "</tr>"
        "</tbody></table></div>"
    )


# --- XLSX ------------------------------------------------------------------
def gerar_xlsx(resultado):
    wb = Workbook()
    ws = wb.active
    ws.title = "Missão COMAVE"

    borda = Border(*[Side(style="thin", color="000000")] * 4)
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    esquerda = Alignment(horizontal="left", vertical="center", wrap_text=True)
    negrito = Font(bold=True)
    vermelho = Font(bold=True, color="FF0000")

    ws.merge_cells("A1:J1")
    ws["A1"] = f"{resultado['contrato']} · SEI {resultado['sei']} · vigência até {resultado['vigencia']}"
    ws["A1"].fill = PatternFill("solid", fgColor="C6E0B4")

    ws.merge_cells("A2:C2")
    ws.merge_cells("D2:F2")
    ws.merge_cells("G2:J2")
    ws["A2"], ws["D2"], ws["G2"] = "ORIGEM", "DESTINO", resultado["titulo_frota"]

    ws.append([
        "MUNICÍPIO - DECOLAGEM", "UF", "ICAO", "MUNICÍPIO - POUSO", "UF", "ICAO",
        "ANV", "TEMPO DE DESLOCAMENTO", "CUSTO ESTIMADO", "CAPACIDADE",
    ])

    for linha in resultado["linhas"]:
        ws.append([
            linha["mun_orig"], linha["uf_orig"], linha["ind_orig"],
            linha["mun_dest"], linha["uf_dest"], linha["ind_dest"],
            linha["anv"], linha["tempo_txt"],
            linha["custo_txt"] if linha["bloqueado"] else linha["custo_num"],
            linha["pax"],
        ])

    linha_total = ws.max_row + 1
    ws.merge_cells(start_row=linha_total, start_column=1, end_row=linha_total, end_column=8)
    ws.cell(row=linha_total, column=1, value="TOTAL")
    ws.merge_cells(start_row=linha_total, start_column=9, end_row=linha_total, end_column=10)
    celula_total = ws.cell(row=linha_total, column=9)
    if resultado["total_bloqueado"]:
        celula_total.value = MSG_TOTAL_BLOQUEADO
        celula_total.font = vermelho
    else:
        celula_total.value = resultado["custo_total"]
        celula_total.number_format = 'R$ #,##0.00'

    linha_nota = linha_total + 1
    ws.merge_cells(start_row=linha_nota, start_column=1, end_row=linha_nota, end_column=10)
    ws.cell(row=linha_nota, column=1, value=NOTA_SOLO).alignment = esquerda
    ws.cell(row=linha_nota, column=1).font = Font(size=9, italic=True)

    linha_fonte = linha_nota + 1
    ws.merge_cells(start_row=linha_fonte, start_column=1, end_row=linha_fonte, end_column=10)
    ws.cell(row=linha_fonte, column=1, value=f"Fonte: {REFERENCIA_NOTA_TECNICA}").alignment = esquerda
    ws.cell(row=linha_fonte, column=1).font = Font(size=9, italic=True)

    for linha_cels in ws.iter_rows(min_row=1, max_row=linha_total, max_col=10):
        for celula in linha_cels:
            celula.border = borda
            celula.alignment = centro

    for col in range(1, 11):
        ws.cell(row=1, column=col).font = negrito
        ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor="9BC2E6")
        ws.cell(row=2, column=col).font = negrito
        ws.cell(row=3, column=col).fill = PatternFill("solid", fgColor="DDEBF7")
        ws.cell(row=3, column=col).font = negrito
        ws.cell(row=linha_total, column=col).fill = PatternFill("solid", fgColor="F2F2F2")
        ws.cell(row=linha_total, column=col).font = negrito

    for idx_linha, linha in enumerate(resultado["linhas"], start=4):
        if linha["bloqueado"]:
            ws.cell(row=idx_linha, column=8).font = vermelho
            ws.cell(row=idx_linha, column=9).font = vermelho
        else:
            ws.cell(row=idx_linha, column=9).number_format = 'R$ #,##0.00'

    for i, largura in enumerate([26, 6, 12, 26, 6, 12, 14, 20, 22, 24], start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- INTERFACE -------------------------------------------------------------
def render(aeroportos_ordenados, cidades_extras, formatador_localidade,
           obter_dados_local, obter_coordenada):
    st.caption(
        "Cálculo com as tabelas de custo do COMAVE. Selecione o órgão antes de montar "
        "a missão — o valor da hora de voo muda de contrato para contrato."
    )

    nome_contrato = st.selectbox(
        "Órgão / Contrato", options=list(CONTRATOS.keys()), key="cmv_contrato"
    )
    contrato = CONTRATOS[nome_contrato]
    frota_disponivel = aeronaves_do_contrato(contrato, FROTA_COMAVE)

    col_a, col_b, col_c = st.columns(3)
    col_a.metric("Processo SEI", contrato["sei"])
    col_b.metric("Vigência até", contrato["vigencia"])
    col_c.metric("Aeronaves cobertas", len(frota_disponivel))
    if contrato.get("observacao"):
        st.info(f"ℹ️ {contrato['observacao']}")

    with st.expander("💰 Tabela de custos deste contrato"):
        st.dataframe(
            pd.DataFrame([
                {"Aeronave": nome, "Valor/hora de voo": formatar_brl(contrato["precos"][nome])}
                for nome in frota_disponivel
            ]),
            width="stretch", hide_index=True,
        )
        st.caption(f"Fonte: {REFERENCIA_NOTA_TECNICA}")

    st.divider()

    if "cmv_trechos" not in st.session_state:
        st.session_state.cmv_trechos = [_novo_trecho(frota_disponivel[0])]
    if "cmv_resultado" not in st.session_state:
        st.session_state.cmv_resultado = None

    # Trocar de contrato pode remover aeronaves da lista; a chave do widget
    # precisa cair junto, senão o Streamlit reclama de valor fora das opções.
    for trecho in st.session_state.cmv_trechos:
        if trecho["aeronave"] not in frota_disponivel:
            trecho["aeronave"] = frota_disponivel[0]
            st.session_state.pop(f"cmv_aeronave_{trecho['id']}", None)
            st.session_state.cmv_resultado = None

    for i, trecho in enumerate(st.session_state.cmv_trechos):
        tid = trecho["id"]
        if i > 0:
            st.markdown(
                "<hr style='border: 2px solid #2E7D32; margin-top: 15px; margin-bottom: 25px;'>",
                unsafe_allow_html=True,
            )
        st.markdown(f"**TRECHO {i + 1}**")
        col_aero, col_orig, col_dest, col_del = st.columns([3.5, 3, 3, 0.5])

        with col_aero:
            trecho["aeronave"] = st.selectbox(
                "Aeronave (Equipamento)",
                options=frota_disponivel,
                index=frota_disponivel.index(trecho["aeronave"]),
                key=f"cmv_aeronave_{tid}",
            )

        opcoes = list(aeroportos_ordenados.keys())
        if FROTA_COMAVE[trecho["aeronave"]].get("is_heli"):
            opcoes += list(cidades_extras.keys())
        for campo in ("origem", "destino"):
            if trecho[campo] not in opcoes:
                trecho[campo] = opcoes[0]
                st.session_state.pop(f"cmv_{campo}_{tid}", None)

        with col_orig:
            trecho["origem"] = st.selectbox(
                "Origem", options=opcoes, index=opcoes.index(trecho["origem"]),
                format_func=formatador_localidade, key=f"cmv_origem_{tid}",
            )
        with col_dest:
            trecho["destino"] = st.selectbox(
                "Destino", options=opcoes, index=opcoes.index(trecho["destino"]),
                format_func=formatador_localidade, key=f"cmv_destino_{tid}",
            )
        with col_del:
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            if len(st.session_state.cmv_trechos) > 1:
                st.button("❌", key=f"cmv_del_{tid}", on_click=_remover_trecho, args=(i,),
                          help="Remover este trecho")

    st.write("")
    col_add, _, col_clear = st.columns([2, 8, 2])
    col_add.button("➕ Adicionar Trecho", key="cmv_add", on_click=_adicionar_trecho)
    col_clear.button("🗑️ Limpar Busca", key="cmv_clear", on_click=_limpar,
                     args=(frota_disponivel[0],), type="secondary", width="stretch")

    st.divider()

    if st.button("Calcular Missão COMAVE", key="cmv_calc", type="primary", width="stretch"):
        st.session_state.cmv_resultado = calcular_missao_comave(
            st.session_state.cmv_trechos, contrato, nome_contrato,
            obter_dados_local, obter_coordenada,
        )

    resultado = st.session_state.cmv_resultado
    if not resultado:
        return

    if resultado["erros"]:
        for erro in resultado["erros"]:
            st.error(f"🚨 {erro}")
        return
    if not resultado["linhas"]:
        return

    if resultado["total_bloqueado"]:
        st.warning(
            "⚠️ Há trecho não operacional pela ANAC (Agência Nacional de Aviação Civil). "
            "A tabela foi gerada, mas o custo total está bloqueado até que o planejamento "
            "seja ajustado."
        )
    else:
        st.success(f"✅ Missão calculada pela tabela do contrato: {resultado['contrato']}.")

    if XLSX_DISPONIVEL:
        carimbo = resultado["gerado_em"].strftime("%Y%m%d_%H%M")
        st.download_button(
            "⬇️ Baixar tabela em Excel (.xlsx)",
            data=gerar_xlsx(resultado),
            file_name=f"missao_comave_{carimbo}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="cmv_download",
        )

    st.markdown(montar_tabela_html(resultado), unsafe_allow_html=True)

    with st.expander("🔍 Memória de cálculo (conferência)"):
        st.dataframe(
            pd.DataFrame([
                {
                    "Trecho": f"{l['ind_orig']} → {l['ind_dest']}",
                    "ANV": l["anv"],
                    "Distância (NM)": l["dist_nm"],
                    "Velocidade (KT)": "Tabelado" if l["tempo_fixo"] else str(l["vel_kt"]),
                    "Tempo de voo": l["tempo_voo_txt"],
                    "Solo (min)": l["minutos_solo"],
                    "Tempo total": l["tempo_txt"],
                    "Horas faturadas": round(l["horas_faturadas"], 4),
                    "Valor/hora (R$)": round(l["valor_hora"], 2),
                    "Custo (R$)": round(l["custo_num"], 2),
                }
                for l in resultado["linhas"]
            ]),
            width="stretch", hide_index=True,
        )
        st.caption(NOTA_SOLO)
